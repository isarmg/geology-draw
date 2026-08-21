"""表格数据读取：支持 Excel（.xlsx/.xlsm）与 CSV。

统一读成 [{列名: 字符串值}, ...]：
- Excel 自动选取"有数据的那张工作表"（表头含岩性/厚度或钻孔等
  关键列、且有数据行者优先），避免封面/说明页在前时误读为空；
  第一行为表头，数据从第二行起；纵向/横向合并单元格自动按左上角值
  展开；公式读取工作簿中已保存的缓存结果（本程序不计算公式）。
- CSV 依次尝试 UTF-8 / GBK 编码（Excel 另存的 CSV 通常是 GBK）。
"""

import csv
import bisect
import os
import re
import xml.etree.ElementTree as ET
import zipfile


# Excel limits here are deliberately much smaller than the OOXML format's
# theoretical 1,048,576 x 16,384 grid.  A stratigraphy document has at most a
# few thousand layers and a few dozen useful columns; accepting a huge sparse
# dimension only makes openpyxl spend CPU and memory on cells the application
# will never use.
MAX_WORKSHEET_ROWS = 10000
MAX_WORKSHEET_COLUMNS = 256
MAX_WORKSHEET_GRID_CELLS = 500000
MAX_EFFECTIVE_CELLS = 250000
MAX_MERGED_RANGES = 5000
MAX_MERGED_AREA = 250000
MAX_WORKSHEET_XML_BYTES = 32 * 1024 * 1024

_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")


def read_table(path):
    """Read a supported table file and return non-empty rows.

    Only CSV, XLSX and XLSM are accepted.  Keeping the accepted extensions
    explicit is important for callers such as web upload handlers: an
    arbitrary file must not silently be treated as CSV merely because its
    suffix is unknown.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    if ext == ".xls":
        raise ValueError("暂不支持旧版 .xls，请在 Excel 中另存为 .xlsx")
    if ext == ".csv":
        return _read_csv(path)
    raise ValueError("不支持的文件格式，仅支持 .csv、.xlsx 和 .xlsm")


def _read_csv(path):
    last_err = None
    for enc in ("utf-8-sig", "gbk", "utf-8"):
        try:
            with open(path, newline="", encoding=enc) as f:
                rows = []
                for raw in csv.DictReader(f):
                    # DictReader stores surplus fields under a None key and
                    # uses a list as the value.  They are not addressable by a
                    # table header, so ignore them rather than calling strip()
                    # on the list.  Normalising all retained values also makes
                    # this helper robust to programmatically supplied values.
                    row = {name: _cell_str(value)
                           for name, value in raw.items() if name is not None}
                    if any(value.strip() for value in row.values()):
                        rows.append(row)
                return rows
        except UnicodeDecodeError as e:
            last_err = e
    raise last_err


def _cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # 8.0 -> "8"，避免层号/厚度出现多余小数
    return str(v).strip()


# 数据表头应含的关键列（任一即可视为数据表）
_KEY_HEADERS = ("岩性", "厚度", "层厚", "钻孔", "层号", "孔口标高")


def _column_number(label):
    value = 0
    for character in label.upper():
        value = value * 26 + ord(character) - ord("A") + 1
    return value


def _cell_bounds(reference):
    """Return ``(min_col, min_row, max_col, max_row)`` for an A1 range."""
    parts = str(reference or "").split(":")
    if len(parts) not in (1, 2):
        raise ValueError("Excel 工作表包含无效的单元格范围")
    matches = [_CELL_REF_RE.match(part) for part in parts]
    if any(match is None for match in matches):
        raise ValueError("Excel 工作表包含无效的单元格范围")
    if len(matches) == 1:
        matches.append(matches[0])
    col1, row1 = _column_number(matches[0].group(1)), int(matches[0].group(2))
    col2, row2 = _column_number(matches[1].group(1)), int(matches[1].group(2))
    if col2 < col1 or row2 < row1:
        raise ValueError("Excel 工作表包含无效的单元格范围")
    return col1, row1, col2, row2


def _validate_bounds(max_row, max_column):
    if max_row > MAX_WORKSHEET_ROWS or max_column > MAX_WORKSHEET_COLUMNS:
        raise ValueError(
            "Excel 工作表尺寸过大（最多 %d 行、%d 列）" %
            (MAX_WORKSHEET_ROWS, MAX_WORKSHEET_COLUMNS))
    if max_row * max_column > MAX_WORKSHEET_GRID_CELLS:
        raise ValueError("Excel 工作表有效区域过大")


def _preflight_xlsx(path):
    """Inspect worksheet XML before openpyxl expands merged-cell objects.

    In normal mode openpyxl materialises every placeholder in a merged range.
    A tiny workbook containing a merge such as ``A1:XFD1048576`` can therefore
    exhaust memory before application-level row limits run.  Streaming the XML
    first lets us reject pathological dimensions without allocating that grid.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            sheets = [info for info in archive.infolist()
                      if (info.filename.startswith("xl/worksheets/")
                          and info.filename.endswith(".xml"))]
            if not sheets:
                raise ValueError("Excel 工作簿不包含可读取的工作表")
            for info in sheets:
                if info.file_size > MAX_WORKSHEET_XML_BYTES:
                    raise ValueError("Excel 工作表数据过大")
                effective_cells = 0
                merge_count = 0
                merged_area = 0
                observed_row = 0
                observed_column = 0
                with archive.open(info) as source:
                    for _event, element in ET.iterparse(source, events=("end",)):
                        tag = element.tag.rsplit("}", 1)[-1]
                        if tag == "dimension":
                            _min_col, _min_row, max_col, max_row = _cell_bounds(
                                element.attrib.get("ref"))
                            _validate_bounds(max_row, max_col)
                        elif tag == "c":
                            effective_cells += 1
                            if effective_cells > MAX_EFFECTIVE_CELLS:
                                raise ValueError("Excel 工作表有效单元格过多")
                            reference = element.attrib.get("r")
                            if reference:
                                col, row, _max_col, _max_row = _cell_bounds(reference)
                                observed_row = max(observed_row, row)
                                observed_column = max(observed_column, col)
                                _validate_bounds(observed_row, observed_column)
                        elif tag == "mergeCell":
                            merge_count += 1
                            if merge_count > MAX_MERGED_RANGES:
                                raise ValueError("Excel 工作表合并区域过多")
                            min_col, min_row, max_col, max_row = _cell_bounds(
                                element.attrib.get("ref"))
                            _validate_bounds(max_row, max_col)
                            merged_area += ((max_row - min_row + 1) *
                                            (max_col - min_col + 1))
                            if merged_area > MAX_MERGED_AREA:
                                raise ValueError("Excel 工作表合并区域过大")
                        element.clear()
    except ValueError:
        raise
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile, ET.ParseError):
        raise ValueError("Excel 工作簿文件无效")


def _read_sheet(ws):
    """读取单张工作表为 [{列名: 值}, ...]（展开合并单元格、跳过空行）。"""
    if ws.max_row < 2 or ws.max_column < 1:
        return [], []
    _validate_bounds(ws.max_row, ws.max_column)

    # Keep one interval per merged range rather than one dictionary entry per
    # covered cell.  Row sweep events make lookup logarithmic while memory is
    # proportional to the number of merge declarations, not their area.
    starts = {}
    ends = {}
    ranges = list(ws.merged_cells.ranges)
    if len(ranges) > MAX_MERGED_RANGES:
        raise ValueError("Excel 工作表合并区域过多")
    merged_area = 0
    for index, rng in enumerate(ranges):
        _validate_bounds(rng.max_row, rng.max_col)
        merged_area += ((rng.max_row - rng.min_row + 1) *
                        (rng.max_col - rng.min_col + 1))
        if merged_area > MAX_MERGED_AREA:
            raise ValueError("Excel 工作表合并区域过大")
        entry = (rng.min_col, rng.max_col,
                 ws.cell(rng.min_row, rng.min_col).value)
        starts.setdefault(rng.min_row, []).append((index, entry))
        ends.setdefault(rng.max_row + 1, []).append(index)

    active = {}

    def row_values(row_number):
        for index in ends.get(row_number, ()):
            active.pop(index, None)
        for index, entry in starts.get(row_number, ()):
            active[index] = entry
        intervals = sorted(active.values(), key=lambda item: item[0])
        interval_starts = [item[0] for item in intervals]

        def val(column_number):
            position = bisect.bisect_right(interval_starts, column_number) - 1
            if position >= 0 and column_number <= intervals[position][1]:
                value = intervals[position][2]
            else:
                value = ws.cell(row_number, column_number).value
            return _cell_str(value)

        return val

    header_value = row_values(1)
    headers = [header_value(c) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        val = row_values(r)
        row = {h: val(c) for c, h in enumerate(headers, 1) if h}
        if any(row.values()):
            rows.append(row)
    return headers, rows


def _read_xlsx(path):
    from openpyxl import load_workbook

    _preflight_xlsx(path)
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        # 逐张工作表评估，优先选"表头含关键列且有数据行"的那张，避免
        # 封面/说明页排在最前时误读为空表。
        best, best_score = [], (-1, -1)
        for ws in wb.worksheets:
            headers, rows = _read_sheet(ws)
            if not rows:
                continue
            has_key = any(any(
                h and h.strip().split("(")[0].split("（")[0] == k
                for h in headers) for k in _KEY_HEADERS)
            score = (1 if has_key else 0, len(rows))
            if score > best_score:
                best, best_score = rows, score
        return best
    finally:
        wb.close()


def get(row, *keys, default=""):
    """按列名（忽略括号后缀，如"厚度(m)"匹配"厚度"）取值。"""
    for k in keys:
        for name, v in row.items():
            if (name and str(name).strip().split("(")[0].split("（")[0]
                    == k):
                return _cell_str(v)
    return default
