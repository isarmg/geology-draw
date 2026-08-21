"""Regression tests for workbook and HTTP resource/security boundaries."""

import http.client
import json
import os
import socket
import tempfile
import threading
import time
import unittest
import zipfile
from pathlib import Path
from unittest import mock


os.environ.setdefault("MPLCONFIGDIR", tempfile.gettempdir())

from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook

from server import APIError, DocumentRepository, create_server
from server import _render_bytes
from strat.tableio import MAX_WORKSHEET_ROWS, read_table


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class WorkbookHardeningTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory(
            prefix="strat-workbook-hardening-")
        self.directory = Path(self.temporary_directory.name)

    def tearDown(self):
        self.temporary_directory.cleanup()

    def _workbook(self, merged=False):
        path = self.directory / "input.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["岩性", "厚度"])
        sheet.append(["砂岩", 1])
        sheet.append([None, 2])
        if merged:
            sheet.merge_cells("A2:A3")
        workbook.save(path)
        workbook.close()
        return path

    @staticmethod
    def _replace_sheet_xml(path, old, new):
        replacement = path.with_name("replacement.xlsx")
        with zipfile.ZipFile(path, "r") as source:
            with zipfile.ZipFile(replacement, "w") as target:
                for info in source.infolist():
                    data = source.read(info.filename)
                    if info.filename == "xl/worksheets/sheet1.xml":
                        if old not in data:
                            raise AssertionError("worksheet fixture did not match")
                        data = data.replace(old, new, 1)
                    target.writestr(info, data)
        os.replace(replacement, path)

    def test_regular_merged_cells_are_still_propagated(self):
        rows = read_table(self._workbook(merged=True))

        self.assertEqual(len(rows), 2)
        self.assertEqual([row["岩性"] for row in rows], ["砂岩", "砂岩"])
        self.assertEqual([row["厚度"] for row in rows], ["1", "2"])

    def test_rejects_giant_merge_before_openpyxl_materialises_it(self):
        path = self._workbook(merged=True)
        self._replace_sheet_xml(
            path,
            b'<mergeCell ref="A2:A3"/>',
            b'<mergeCell ref="A2:XFD1048576"/>',
        )

        with mock.patch("openpyxl.load_workbook") as loader:
            with self.assertRaisesRegex(ValueError, "合并区域|尺寸过大"):
                read_table(path)
        loader.assert_not_called()

    def test_rejects_pathological_declared_dimension(self):
        path = self._workbook()
        self._replace_sheet_xml(
            path,
            b'<dimension ref="A1:B3"/>',
            b'<dimension ref="A1:XFD1048576"/>',
        )

        with self.assertRaisesRegex(ValueError, "尺寸过大"):
            read_table(path)

    def test_workbook_is_closed_when_sheet_validation_fails(self):
        path = self._workbook()
        workbook = load_workbook(path)
        workbook.active.cell(MAX_WORKSHEET_ROWS + 1, 1, "too far")
        original_close = workbook.close
        workbook.close = mock.Mock(wraps=original_close)

        with mock.patch("openpyxl.load_workbook", return_value=workbook):
            with self.assertRaisesRegex(ValueError, "尺寸过大"):
                read_table(path)

        workbook.close.assert_called_once_with()


class HTTPHardeningTests(unittest.TestCase):
    def setUp(self):
        self.servers = []

    def tearDown(self):
        for server, thread in reversed(self.servers):
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)

    def _start(self, **options):
        server = create_server(
            "127.0.0.1", 0,
            web_dir=str(PROJECT_ROOT / "web"),
            example_dir=str(PROJECT_ROOT / "examples"),
            **options
        )
        server.RequestHandlerClass.log_message = lambda *_args: None
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        self.servers.append((server, thread))
        return server

    @staticmethod
    def _request(server, method="GET", path="/api/v1/health", headers=None,
                 timeout=3):
        host, port = server.server_address[:2]
        connection = http.client.HTTPConnection(host, port, timeout=timeout)
        try:
            connection.request(method, path, headers=dict(headers or {}))
            response = connection.getresponse()
            return response.status, dict(response.getheaders()), response.read()
        finally:
            connection.close()

    def test_loopback_mode_rejects_dns_rebinding_host(self):
        server = self._start()

        status, _headers, body = self._request(
            server, headers={"Host": "attacker.example"})

        self.assertEqual(status, 403)
        self.assertEqual(json.loads(body)["error"]["code"], "host_not_allowed")

    def test_origin_must_match_request_host(self):
        server = self._start()
        host, port = server.server_address[:2]
        authority = "%s:%d" % (host, port)

        denied = self._request(server, headers={
            "Host": authority,
            "Origin": "http://attacker.example",
        })
        allowed = self._request(server, headers={
            "Host": authority,
            "Origin": "http://" + authority,
        })

        self.assertEqual(denied[0], 403)
        self.assertEqual(json.loads(denied[2])["error"]["code"],
                         "origin_not_allowed")
        self.assertEqual(allowed[0], 200)

    def test_non_loopback_bind_requires_explicit_opt_in(self):
        with self.assertRaisesRegex(ValueError, "allow_remote|allow-network"):
            create_server(
                "0.0.0.0", 0,
                web_dir=str(PROJECT_ROOT / "web"),
                example_dir=str(PROJECT_ROOT / "examples"),
            )

        server = create_server(
            "0.0.0.0", 0,
            web_dir=str(PROJECT_ROOT / "web"),
            example_dir=str(PROJECT_ROOT / "examples"),
            allow_remote=True,
        )
        server.server_close()

    def test_partial_body_times_out(self):
        server = self._start(request_timeout=0.2)
        host, port = server.server_address[:2]
        client = socket.create_connection((host, port), timeout=2)
        client.settimeout(2)
        try:
            request = (
                "POST /api/v1/examples/column HTTP/1.1\r\n"
                "Host: %s:%d\r\n"
                "Content-Type: application/json\r\n"
                "Content-Length: 2\r\n\r\n{" % (host, port)
            ).encode("ascii")
            client.sendall(request)
            chunks = []
            while True:
                chunk = client.recv(4096)
                if not chunk:
                    break
                chunks.append(chunk)
            response = b"".join(chunks)
        finally:
            client.close()

        self.assertIn(b" 408 ", response)
        self.assertIn(b"request_timeout", response)

    def test_active_request_threads_are_bounded(self):
        server = self._start(max_workers=1, request_timeout=2)
        host, port = server.server_address[:2]
        blocker = socket.create_connection((host, port), timeout=2)
        try:
            blocker.sendall((
                "POST /api/v1/examples/column HTTP/1.1\r\n"
                "Host: %s:%d\r\n"
                "Content-Type: application/json\r\n"
                "Content-Length: 2\r\n\r\n" % (host, port)
            ).encode("ascii"))
            deadline = time.monotonic() + 1
            while (getattr(server._request_slots, "_value", 1) != 0
                   and time.monotonic() < deadline):
                time.sleep(0.01)

            response = self._request(server)
        finally:
            blocker.close()

        self.assertEqual(response[0], 503)
        self.assertEqual(json.loads(response[2])["error"]["code"],
                         "server_busy")


class RenderMemoryBoundaryTests(unittest.TestCase):
    def test_oversized_canvas_is_rejected_before_savefig(self):
        figure = Figure(figsize=(100, 100))

        with self.assertRaises(APIError) as raised:
            _render_bytes(lambda _data: figure, None, None, None, 8.0,
                          "png", 600)

        self.assertEqual(raised.exception.code, "render_too_large")


class RepositoryMemoryBoundaryTests(unittest.TestCase):
    @staticmethod
    def _data(description=""):
        return [{"lith": "砂岩", "thick": 1.0, "desc": description}]

    def test_repository_enforces_byte_quota_and_delete_releases_it(self):
        sizing = DocumentRepository(max_bytes=10 ** 7)
        sized = sizing.add("sizing.csv", "column", self._data("x" * 200))
        repository = DocumentRepository(
            max_documents=10, max_bytes=sized.size_bytes + 16)

        first = repository.add("first.csv", "column", self._data("x" * 200))
        self.assertEqual(repository.bytes_used, first.size_bytes)
        with self.assertRaises(APIError) as raised:
            repository.add("second.csv", "column", self._data("x" * 200))
        self.assertEqual(raised.exception.code, "repository_full")

        repository.delete(first.id)
        self.assertEqual(repository.bytes_used, 0)
        repository.add("replacement.csv", "column", self._data("x" * 200))

    def test_expired_documents_release_byte_quota(self):
        now = [0.0]
        repository = DocumentRepository(
            ttl_seconds=1, max_bytes=10 ** 7, clock=lambda: now[0])
        repository.add("old.csv", "column", self._data("old"))
        self.assertGreater(repository.bytes_used, 0)

        now[0] = 2.0
        self.assertEqual(len(repository), 0)
        self.assertEqual(repository.bytes_used, 0)


if __name__ == "__main__":
    unittest.main()
